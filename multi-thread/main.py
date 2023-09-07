from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage,filedialog,END,ttk,Variable,messagebox
from tkinter.filedialog import askopenfilename
from pathlib import Path
import win32api
import win32con
import openpyxl
import time
import cv2
import pygame
import numpy as np
import os
import threading
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.chrome.options import Options
from undetected_chromedriver import Chrome, ChromeOptions
from io import BytesIO
from PIL import Image

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path("assets/")

pygame.init()
address_btn_xpath = '//*[@id="intent-banner-section"]/button/div/div[2]'
address_add_btn_xpath = '//*[@id="intent-banner-section"]/section[1]/div/div/div/div[1]/div/div[1]/div/div/div/div[2]/button'
address_select_btn_xpath = '//*[@id="intent-banner-section"]/section[1]/div/div/div/div[1]/div/div[1]/div/button'
address_edit_btn_xpath = '//*[@id="address-selection-form"]/ul/li/div/div/button'

change_form_xpath = '//*[@id="add-edit-address-form2"]'

firstname_input_xpath = '//*[@id="react-aria-7"]'
lastname_input_xpath = '//*[@id="react-aria-8"]'
street_input_xpath = '//*[@id="addressLineOne"]'
city_input_xpath = '//*[@id="react-aria-10"]'
state_input_xpath = '//*[@id="react-aria-11"]'
zipcode_input_xpath = '//*[@id="react-aria-12"]'
phonenumber_input_xpath ='//*[@id="react-aria-13"]'
saveaddress_btn_xpath = '/html/body/div[2]/div/div[2]/div[1]/div/div[3]/button'
useaddress_btn_xpath = '/html/body/div[2]/div/div[2]/div[1]/div/div[2]/div/div[2]/div/button'
selectaddress_btn_xpath = '//*[@id="address-selection-form"]/ul/li'
selectsave_btn_xpath = '/html/body/div[2]/div/div[2]/div[1]/div/div[3]/button'
exitediting_btn_xpath = '/html/body/div[2]/div/div[2]/div[1]/div/div[1]/button'

stop_event = threading.Event()

wait_time = 0
remaining_time = 0
address1_value = ""
city1_value = ""
state1_value = ""
telephone1_value = ""
zipCode1_value = ""
address2_value = ""
city2_value = ""
state2_value = ""
telephone2_value = ""
zipCode2_value = ""
address3_value = ""
city3_value = ""
state3_value = ""
telephone3_value = ""
zipCode3_value = ""
pixelRatio = 1

num_errors = 0
array_errors = []
scanning_initial = True

# Create a lock object
lock = threading.Lock()

window = Tk()
window.title("Crawler")
window.geometry("800x680")
window.configure(bg = "#202020")
canvas = Canvas(
    window,
    bg = "#202020",
    height = 680,
    width = 800,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

# Play an audio file using pygame
def play_audio(filename):
    pygame.mixer.music.load(filename)
    pygame.mixer.music.play()

def add_address(street, city, state, telephone, zipCode, driver, wait):
    change_addressForm(street, city, state, telephone, zipCode, 0, driver, wait)

    try:
        selectaddress_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, selectaddress_btn_xpath)))
        selectaddress_btn.click()
    except:
        useaddress_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, useaddress_btn_xpath)))
        useaddress_btn.click()
        selectaddress_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, selectaddress_btn_xpath)))
        selectaddress_btn.click()

    selectsave_btn = driver.find_element(By.XPATH, selectsave_btn_xpath)
    selectsave_btn.click()
    
    time.sleep(2)
    address_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
    address_btn.click()


def change_addressForm(street, city, state, telephone, zipCode, status, driver, wait):
    change_form_element = wait.until(EC.presence_of_element_located(
        (By.XPATH, change_form_xpath)))

    div_elements = change_form_element.find_elements(By.XPATH, './div')

    if(status == 0):
        firstname_field = div_elements[0].find_element(By.TAG_NAME, 'input')
        firstname_field.send_keys(Keys.CONTROL + 'a')  # Select all text
        firstname_field.send_keys(Keys.DELETE)         # Delete selected text
        firstname_field.send_keys("A")

        lastname_field = div_elements[1].find_element(By.TAG_NAME, 'input')
        lastname_field.send_keys(Keys.CONTROL + 'a')  # Select all text
        lastname_field.send_keys(Keys.DELETE)         # Delete selected text
        lastname_field.send_keys("B")

    street_field = div_elements[2].find_element(By.TAG_NAME, 'input')
    street_field.send_keys(Keys.CONTROL + 'a')  # Select all text
    street_field.send_keys(Keys.DELETE)         # Delete selected text
    street_field.send_keys(street)

    city_field = div_elements[4].find_element(By.TAG_NAME, 'input')
    city_field.send_keys(Keys.CONTROL + 'a')  # Select all text
    city_field.send_keys(Keys.DELETE)         # Delete selected text
    city_field.send_keys(city)


    state_field = div_elements[5].find_element(By.TAG_NAME, 'select')
    states_select = Select(state_field)
    states_select.select_by_value(state)

    zipcode_field = div_elements[5].find_element(By.TAG_NAME, 'input')
    zipcode_field.send_keys(Keys.CONTROL + 'a')  # Select all text
    zipcode_field.send_keys(Keys.DELETE)         # Delete selected text
    zipcode_field.send_keys(zipCode)

    phonenumber_field = div_elements[6].find_element(By.TAG_NAME, 'input')
    phonenumber_field.send_keys(Keys.CONTROL + 'a')  # Select all text
    phonenumber_field.send_keys(Keys.DELETE)         # Delete selected text
    phonenumber_field.send_keys(telephone)

    saveaddress_btn = driver.find_element(By.XPATH, saveaddress_btn_xpath)
    saveaddress_btn.click()
    
def editAddress(street, city, state, telephone, zipCode, driver, wait):
    address_select_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_select_btn_xpath)))
    address_select_btn.click()

    address_edit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_edit_btn_xpath)))
    address_edit_btn.click()

    change_addressForm(street, city, state, telephone, zipCode, 1, driver, wait)

    try:
        selectaddress_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, selectaddress_btn_xpath)))
        selectaddress_btn.click()
    except:
        useaddress_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, useaddress_btn_xpath)))
        useaddress_btn.click()
        selectaddress_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, selectaddress_btn_xpath)))
        selectaddress_btn.click()


    selectsave_btn = driver.find_element(By.XPATH, selectsave_btn_xpath)
    selectsave_btn.click()
    time.sleep(2)
    address_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
    address_btn.click()

def check_compare_sellers(driver, wait):
    try:
        compare_sellers_btn = driver.find_element(By.CSS_SELECTOR, 'button[aria-label="Compare all sellers"]')
        compare_sellers_btn.click()
        
        shipping_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="allSellersOfferLine"]')))
        if(shipping_element.find_element(By.CSS_SELECTOR, '.lh-copy').get_attribute('innerText') == 'Sold and shipped by Walmart.com'):
            close_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div/div[2]/div[1]/div/div[1]/button')
            close_btn.click()
            return 1
        else:
            close_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div/div[2]/div[1]/div/div[1]/button')
            close_btn.click()
            return 0

    except:
        return 0

def check_shipping(driver, wait):
    try:
        sold_shipped_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="sold-and-shipped-by"]')))
        if(sold_shipped_element.find_element(By.CSS_SELECTOR, '.lh-copy').get_attribute('innerText') == 'Sold and shipped by Walmart.com'):
            return 1
        else:
            check_compare_sellers(driver, wait)
    except:
        check_compare_sellers(driver, wait)
            
    return 0

def check_status(link, driver, wait):
    if stop_event.is_set():
        print("Thread stopped.")
        return "stopped"
    
    if(check_shipping(driver, wait) == 1):
        time.sleep(2)
        addcart_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="add-to-cart-section"]'))).find_element(By.TAG_NAME, 'button')
        addcart_btn.click()

        # stime.sleep(1)
        action_chains = ActionChains(driver)
        cartview_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="cart-button-header"]')))
        # cartview_btn.click()
        action_chains.double_click(cartview_btn).perform()

        try:
            shipping_element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="maincontent"]/div/div/div/div[2]/div[1]/div[2]/div/button[1]')))
            shipping_label = shipping_element.get_attribute('aria-label')
        except:
            try:
                dropdown_btn = driver.find_element(By.XPATH, '//*[@id="maincontent"]/div/div/div/div[2]/div[1]/div[2]/button')
                dropdown_btn.click()
                shipping_element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="maincontent"]/div/div/div/div[2]/div[1]/div[2]/div/button[1]')))
                shipping_label = shipping_element.get_attribute('aria-label')
            except:
                driver.get(link)
                return 0

        try:
            if shipping_label.split(",")[1].strip() == 'Available' or shipping_label.split(",")[1].strip() == 'All Items Available':
                remove_btn = driver.find_element(By.XPATH, '//*[@id="maincontent"]/div/div/div/div[2]/div[1]/div[3]/div/div/div[2]/section/div/div/div/ul/li/div[3]/div[2]/section/button[1]')
                remove_btn.click()
                driver.get(link)
                return 1
            else:
                remove_btn = driver.find_element(By.XPATH, '//*[@id="maincontent"]/div/div/div/div[2]/div[1]/div[3]/div/div/div[2]/section/div/div/div/ul/li/div[3]/div[2]/section/button[1]')
                remove_btn.click()
                driver.get(link)
                return 0
        except:
            driver.get(link)
            return 0
    else:
        return 0
    
def solve_blocked(driver, retry, link, index, wait, thread_id):
    global captcha_status
    '''
    Solve blocked
    (Cross-domain iframe cannot get elements temporarily)
    Simulate the mouse press and hold to complete the verification
    '''
    element = None
    try:
        element = WebDriverWait(driver,1).until(EC.element_to_be_clickable((By.ID,'px-captcha')))
    except BaseException as e:
        print(f'px-captcha element not found')
        captcha_status = "true"
        return
    
    if not retry:
        driver.quit()
        captcha_status = "false"
        return
    
    print(f'solve blocked:{driver.current_url}, Retry {retry} remaining times')
    template = cv2.imread(os.path.join('./captcha.png'), 0)
    # Set the minimum number of feature points to match value 10
    MIN_MATCH_COUNT = 10
    if  element:
        print(f'start press and hold')
        x, y = element.location['x'], element.location['y']
        width, height = element.size.get('width'), element.size.get('height')        
        x_move = x + width * 0.5
        y_move = y + height * 0.5

        ActionChains(driver).move_to_element_with_offset(element, 0, 0).click_and_hold().perform()
        time.sleep(1)
        # ActionChains(driver).moveToElement(element).build().perform()
        # ActionChains(driver).click_and_hold(element).perform()
        start_time = time.time()
        while 1:
            if time.time() - start_time > 15:
                break
            left = x * pixelRatio
            top = y * pixelRatio
            right = (x+width) * pixelRatio
            bottom = (y+height) * pixelRatio
            png = driver.get_screenshot_as_png() 
            im = Image.open(BytesIO(png))
            im = im.crop((left, top, right, bottom))
            target = cv2.cvtColor(np.asarray(im),cv2.COLOR_RGB2BGR)  
            # Initiate SIFT detector
            sift = cv2.SIFT_create()
            # find the keypoints and descriptors with SIFT
            kp1, des1 = sift.detectAndCompute(template,None)
            kp2, des2 = sift.detectAndCompute(target,None)
            
            # create set FLANN match
            FLANN_INDEX_KDTREE = 0
            index_params = dict(algorithm = FLANN_INDEX_KDTREE, trees = 5)
            search_params = dict(checks = 50)
            flann = cv2.FlannBasedMatcher(index_params, search_params)
            matches = flann.knnMatch(des1,des2,k=2)
            # store all the good matches as per Lowe's ratio test.
            good = []

            # Discard matches greater than 0.7
            for match in matches:
                if len(match) == 2:
                    m, n = match
                    if m.distance < 0.7*n.distance:
                        good.append(m)

            print( "Not enough matches are found - %d/%d" % (len(good),MIN_MATCH_COUNT))
            if len(good)>=MIN_MATCH_COUNT:
                print(f'release button')
                ActionChains(driver).release(element).perform()
                time.sleep(5)
                break
            time.sleep(0.2)
                 
    time.sleep(5)
    retry -= 1
    solve_blocked(driver, retry, link, index, wait, thread_id)

def check_product(link, index, driver, thread_id, sub_index):    
    global address1_value, address2_value, address3_value
    global city1_value, city2_value, city3_value
    global state1_value, state2_value, state3_value
    global telephone1_value, telephone2_value, telephone3_value
    global zipCode1_value, zipCode2_value, zipCode3_value
    global captcha_status
    
    wait = WebDriverWait(driver, 30)

    try:
        try:
            driver.get(link)
        except Exception as e:
            return "An error occurred:"+ str(e)

        # Check the status code of the page
        status_code = driver.execute_script('return document.status')
        if status_code == 404:
            return "404 error"
        
        solve_blocked(driver, 3, link, index, wait, thread_id)
        if(captcha_status == "false"):
            return "captcha false"
        
        flag = 0
        
        if thread_id == 0:
            # Clear the current value
            productURL1_entry.delete(0, "end")
            productURL1_entry.insert(0, link)
        elif thread_id == 1:
            # Clear the current value
            productURL2_entry.delete(0, "end")
            productURL2_entry.insert(0, link)
        elif thread_id == 2:
            # Clear the current value
            productURL3_entry.delete(0, "end")
            productURL3_entry.insert(0, link)
        elif thread_id == 3:
            # Clear the current value
            productURL4_entry.delete(0, "end")
            productURL4_entry.insert(0, link)

        print(str(thread_id) + str(index))
        
        #add address
        if(index-sub_index == 0):
            address_btn = WebDriverWait(driver, 150).until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
            address_btn.click()
            try:
                address_add_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, address_add_btn_xpath)))
                address_add_btn.click()
                add_address(address2_value, city2_value, state2_value, telephone2_value, zipCode2_value, driver, wait)
            except:
                print("no address add button")

        if((index-sub_index) % 2 == 0):
            #address 1
            try:
                status = check_status(link, driver, wait)
                if(status == "stopped"):
                    return "stopped"
                flag += status
            except Exception as e:
                flag += 0
                print("An error occurred:"+ str(e))
            
            #address2
            if address2_value and city2_value and state2_value and telephone2_value and zipCode2_value:
                try:
                    address_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
                    address_btn.click()
                    try:
                        address_add_btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, address_add_btn_xpath)))
                        address_add_btn.click()
                        add_address(address2_value, city2_value, state2_value, telephone2_value, zipCode2_value, driver, wait)
                    except:
                        print("no address add button")
                        editAddress(address2_value, city2_value, state2_value, telephone2_value, zipCode2_value, driver, wait)
                    status = check_status(link, driver, wait)
                    if(status == "stopped"):
                        return "stopped"
                    flag += status
                except Exception as e:
                    flag += 0
                    print("An error occurred:"+ str(e))
                
            #address 3
            if address3_value and city3_value and state3_value and telephone3_value and zipCode3_value:
                try:
                    address_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
                    address_btn.click()
                    try:
                        address_add_btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, address_add_btn_xpath)))
                        address_add_btn.click()
                        add_address(address3_value, city3_value, state3_value, telephone3_value, zipCode3_value, driver, wait)
                    except:
                        print("no address add button")
                        editAddress(address3_value, city3_value, state3_value, telephone3_value, zipCode3_value, driver, wait)
                    status = check_status(link, driver, wait)
                    if(status == "stopped"):
                        return "stopped"
                    flag += status
                except Exception as e:
                    flag += 0
                    print("An error occurred:"+ str(e))
        else:
            #address 3
            try:
                status = check_status(link, driver, wait)
                if(status == "stopped"):
                    return "stopped"
                flag += status
            except Exception as e:
                flag += 0
                print("An error occurred:"+ str(e))

            #address2
            if address2_value and city2_value and state2_value and telephone2_value and zipCode2_value:
                try:
                    address_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
                    address_btn.click()
                    try:
                        address_add_btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, address_add_btn_xpath)))
                        address_add_btn.click()
                        add_address(address2_value, city2_value, state2_value, telephone2_value, zipCode2_value, driver, wait)
                    except:
                        print("no address add button")
                        editAddress(address2_value, city2_value, state2_value, telephone2_value, zipCode2_value, driver, wait)
                    status = check_status(link, driver, wait)
                    if(status == "stopped"):
                        return "stopped"
                    flag += status
                except Exception as e:
                    flag += 0
                    print("An error occurred:"+ str(e))
            
            #address1
            if address1_value and city1_value and state1_value and telephone1_value and zipCode1_value:
                try:
                    address_btn = wait.until(EC.element_to_be_clickable((By.XPATH, address_btn_xpath)))
                    address_btn.click()
                    try:
                        address_add_btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, address_add_btn_xpath)))
                        address_add_btn.click()
                        add_address(address1_value, city1_value, state1_value, telephone1_value, zipCode1_value, driver, wait)
                    except:
                        print("no address add button")
                        editAddress(address1_value, city1_value, state1_value, telephone1_value, zipCode1_value, driver, wait)
                    status = check_status(link, driver, wait)
                    if(status == "stopped"):
                        return "stopped"
                    flag += status
                except Exception as e:
                    flag += 0
                    print("An error occurred:"+ str(e))
        
        if(flag >= 2):
            return "TRUE"
        else:
            return "FALSE"

    except Exception as e:
        return "An error occurred:"+ str(e)

def run_process(start_row, end_row, scan_index, sheet, thread_id, driver, error_sheet):
    global num_errors
    global array_errors
    global scanning_initial
    
    row_index = start_row
    sub_index = 0

    for i, row in enumerate(sheet.iter_rows(min_row=row_index, max_row = end_row, values_only=True)):
        # Check if the stop event is set
        if stop_event.is_set():
            break
        
        start_row = row_index + i
        if thread_id == 0:
            # Clear the current value
            productID1_entry.delete(0, "end")
            productID1_entry.insert(0, start_row)
        elif thread_id == 1:
            # Clear the current value
            productID2_entry.delete(0, "end")
            productID2_entry.insert(0, start_row)
        elif thread_id == 2:
            # Clear the current value
            productID3_entry.delete(0, "end")
            productID3_entry.insert(0, start_row)
        elif thread_id == 3:
            # Clear the current value
            productID4_entry.delete(0, "end")
            productID4_entry.insert(0, start_row)

        value = check_product(row[0], i, driver, thread_id, sub_index)
        while (value == "captcha false"):
            time.sleep(600)
            try:
                options = ChromeOptions()
                options.add_argument("--disable-extensions")
                options.add_argument("--disable-gpu")
                options.add_argument("--disable-infobars")
                options.add_argument("--disable-notifications")
                options.add_argument("--disable-web-security")
                options.add_argument("--no-sandbox")
                options.add_argument("--disable-dev-shm-usage")
                options.add_argument("--disable-software-rasterizer")
                options.add_argument("--disable-features=VizDisplayCompositor")
                options.add_argument("--blink-settings=imagesEnabled=false")
                options.add_argument("--blink-settings=videoEnabled=false")
                options.add_argument("--disable-css-rendering")
                # or alternatively we can set direct preference:
                prefs = {'profile.default_content_setting_values': {'images': 2, 'css': 2}}
                options.add_experimental_option('prefs', prefs)
                
                driver = Chrome(executable_path=ChromeDriverManager().install(), options=options)

                driver.set_page_load_timeout(30)
                sub_index = i
                value = check_product(row[0], i, driver, thread_id, sub_index)
            except:
                driver = None
                time.sleep(300)
                continue

        if value == "stopped":
            break
        
        if "error" in value:
            with lock:
                num_errors += 1
                array_errors.append(row_index + i)

        try:
            b_value = row[1]

            if(scanning_initial == True):
                sheet.cell(row=row_index + i, column=2).value = value
            else:
                sheet.cell(row=row_index + i, column=3).value = value
                if "error" in value or "error" in b_value:
                    d_value = "error"
                else:
                    if(b_value == value):
                        d_value = "0"
                    else:
                        if(b_value == "TRUE"):
                            d_value = "out"
                        else:
                            d_value = "in"

                sheet.cell(row=row_index + i, column=4).value = d_value
        except:
            sheet.cell(row=row_index + i, column=2).value = value
    
    driver.quit()
    driver = None
    
def update_timer():
    global remaining_time
    # if remaining_time > 0:
    #     text = "Time left: " + str(remaining_time) + " seconds"
    #     canvas.itemconfigure(remainingtime_text, text=text)
    #     remaining_time -= 1
    #     window.after(1000, update_timer)
    # else:
    #     canvas.itemconfigure(remainingtime_text, text="Time is Up!")
    if(stop_event.is_set()):
        canvas.itemconfigure(remainingtime_text, text="")
    else:
        if remaining_time > 0:
            text = "Time left: " + str(remaining_time) + " seconds"
            canvas.itemconfigure(remainingtime_text, text=text)
            remaining_time -= 1
            window.after(1000, update_timer)
        else:
            canvas.itemconfigure(remainingtime_text, text="Time is Up!")

def scan_errors(sheet):
    print("scan error start")
    global num_errors
    global array_errors
    global scanning_initial

    errors = []
    while(1):
        try:
            options = ChromeOptions()
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-web-security")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-software-rasterizer")
            options.add_argument("--disable-features=VizDisplayCompositor")
            options.add_argument("--blink-settings=imagesEnabled=false")
            options.add_argument("--blink-settings=videoEnabled=false")
            options.add_argument("--disable-css-rendering")
            # or alternatively we can set direct preference:
            prefs = {'profile.default_content_setting_values': {'images': 2, 'css': 2}}
            options.add_experimental_option('prefs', prefs)
            
            driver = Chrome(executable_path=ChromeDriverManager().install(), options=options)
            driver.set_page_load_timeout(30)
            break
        except:
            driver = None
            time.sleep(300)
            continue

    sub_index = 0
    index = 0
    num_errors = 0

    for error_index in array_errors:
        link = sheet.cell(row = error_index, column = 1).value
        value = check_product(link, index, driver, 0, sub_index)
        while (value == "captcha false"):
            time.sleep(600)
            try:
                options = ChromeOptions()
                options.add_argument("--disable-extensions")
                options.add_argument("--disable-gpu")
                options.add_argument("--disable-infobars")
                options.add_argument("--disable-notifications")
                options.add_argument("--disable-web-security")
                options.add_argument("--no-sandbox")
                options.add_argument("--disable-dev-shm-usage")
                options.add_argument("--disable-software-rasterizer")
                options.add_argument("--disable-features=VizDisplayCompositor")
                options.add_argument("--blink-settings=imagesEnabled=false")
                options.add_argument("--blink-settings=videoEnabled=false")
                options.add_argument("--disable-css-rendering")
                # or alternatively we can set direct preference:
                prefs = {'profile.default_content_setting_values': {'images': 2, 'css': 2}}
                options.add_experimental_option('prefs', prefs)
                
                driver = Chrome(executable_path=ChromeDriverManager().install(), options=options)

                driver.set_page_load_timeout(30)
                sub_index = index
                value = check_product(link, index, driver, 0, sub_index)
            except:
                driver = None
                time.sleep(300)
                continue

        if value == "stopped":
            break
        
        if "error" in value:
            with lock:
                num_errors += 1
                errors.append(error_index)

        try:
            b_value = sheet.cell(row = error_index, column = 2).value

            if(scanning_initial == True):
                sheet.cell(row=error_index, column=2).value = value
            else:
                sheet.cell(row=error_index, column=3).value = value
                if "error" in value or "error" in b_value:
                    d_value = "error"
                else:
                    if(b_value == value):
                        d_value = "0"
                    else:
                        if(b_value == "TRUE"):
                            d_value = "out"
                        else:
                            d_value = "in"

                sheet.cell(row=error_index, column=4).value = d_value
        except:
            sheet.cell(row=error_index, column=2).value = value

        index = index +1

    driver.quit()
    driver = None
    array_errors = errors

def main_process(workbook, sheet, directory_path):
    global wait_time
    global remaining_time
    global num_errors
    global array_errors
    global scanning_initial

    num_errors = 0
    row_count = sheet.max_row
    
    scan_index = 0
    
    error_workbook = openpyxl.Workbook()
    error_sheet = error_workbook.active
    error_sheet.title = "Sheet1"

    error_sheet['A1'] = "product link"
    error_sheet['B1'] = "could be ordered by walmart's shipping method(first scan)"

    while not stop_event.is_set():
        scan_index += 1
        num_errors = 0
        scanning_text = "Scanning " +  str(scan_index)
        canvas.itemconfigure(scanning_id_text, text=scanning_text)
        canvas.itemconfigure(scanning_status_text, text="running")
        canvas.itemconfigure(remainingtime_text, text="")

        threads = []
        try:
            threads_cnt = int(threadscnt_entry.get())
        except:
            threads_cnt = 1

        count = row_count // threads_cnt
        end_row = 1

        if (sheet.cell(row=2, column=2).value != None):
            scanning_initial = False
        else:
            scanning_initial = True

        for i in range(threads_cnt):
            start_row = end_row
            if(i == threads_cnt - 1):
                end_row = row_count
            else:
                end_row = start_row + count

            driver = None
            while(1):
                try:
                    options = ChromeOptions()
                    options.add_argument("--disable-extensions")
                    options.add_argument("--disable-gpu")
                    options.add_argument("--disable-infobars")
                    options.add_argument("--disable-notifications")
                    options.add_argument("--disable-web-security")
                    options.add_argument("--no-sandbox")
                    options.add_argument("--disable-dev-shm-usage")
                    options.add_argument("--disable-software-rasterizer")
                    options.add_argument("--disable-features=VizDisplayCompositor")
                    options.add_argument("--blink-settings=imagesEnabled=false")
                    options.add_argument("--blink-settings=videoEnabled=false")
                    options.add_argument("--disable-css-rendering")
                    # or alternatively we can set direct preference:
                    prefs = {'profile.default_content_setting_values': {'images': 2, 'css': 2}}
                    options.add_experimental_option('prefs', prefs)
                    
                    driver = Chrome(executable_path=ChromeDriverManager().install(), options=options)
                    driver.set_page_load_timeout(30)
                    break
                except:
                    driver = None
                    time.sleep(300)
                    continue
            
            # options = ChromeOptions()
            # options.add_argument("--disable-extensions")
            # options.add_argument("--disable-gpu")
            # options.add_argument("--disable-infobars")
            # options.add_argument("--disable-notifications")
            # options.add_argument("--disable-web-security")
            # options.add_argument("--no-sandbox")
            # options.add_argument("--disable-dev-shm-usage")
            # options.add_argument("--disable-software-rasterizer")
            # options.add_argument("--disable-features=VizDisplayCompositor")
            # options.add_argument("--blink-settings=imagesEnabled=false")
            # options.add_argument("--blink-settings=videoEnabled=false")
            # options.add_argument("--disable-css-rendering")
            # # or alternatively we can set direct preference:
            # prefs = {'profile.default_content_setting_values': {'images': 2, 'css': 2}}
            # options.add_experimental_option('prefs', prefs)
            
            # driver = Chrome(executable_path=ChromeDriverManager().install(), options=options)
            # driver.set_page_load_timeout(30)
            t = threading.Thread(target=run_process, args=(start_row+1, end_row, scan_index, sheet, i, driver, error_sheet))
            threads.append(t)
            t.start()
            time.sleep(100)
        
        for t in threads:
            t.join()

        threads = None
        running_cnt = 5

        if(stop_event.is_set()):
            # Get the current date and time
            current_datetime = datetime.now()
            current_date = current_datetime.date()
            current_hour = current_datetime.hour
            current_minute = current_datetime.minute
            current_second = current_datetime.second

            #save excel file after scanning all products
            save_filename = directory_path + "/walmart_" + str(current_date) + "_" + str(current_hour) + "_" + str(current_minute) + "_" + str(current_second) + ".xlsx"
            workbook.save(save_filename)
            # error_filename = directory_path + "/errors_" + str(current_date) + "_" + str(current_hour) + "_" + str(current_minute) + "_" + str(current_second) + ".xlsx"
            # error_workbook.save(error_filename)

            canvas.itemconfigure(errorproducts_text, text= str(num_errors) + " errors")
            
            while num_errors > 0 and running_cnt >= 0:
                time.sleep(300)
                scan_errors(sheet)
                running_cnt = running_cnt -1
                workbook.save(save_filename)
                canvas.itemconfigure(errorproducts_text, text= str(num_errors) + " errors")

            try:
                # Display a pop-up notification and play audio
                pygame.mixer.init()
                pygame.mixer.music.load(relative_to_assets('notification.mp3'))
                pygame.mixer.music.play()
            except:
                print("audio error")

            messagebox_text = "Scanning Stopped"
            # Display a pop-up notification
            win32api.MessageBox(
                None,
                messagebox_text,
                "Notification",
                win32con.MB_OK | win32con.MB_ICONINFORMATION
            )
        else:
            # Get the current date and time
            current_datetime = datetime.now()
            current_date = current_datetime.date()
            current_hour = current_datetime.hour
            current_minute = current_datetime.minute
            current_second = current_datetime.second

            #save excel file after scanning all products
            save_filename = directory_path + "/walmart_" + str(current_date) + "_" + str(current_hour) + "_" + str(current_minute) + "_" + str(current_second) + ".xlsx"
            workbook.save(save_filename)
            # error_filename = directory_path + "/errors_" + str(current_date) + "_" + str(current_hour) + "_" + str(current_minute) + "_" + str(current_second) + ".xlsx"
            # error_workbook.save(error_filename)
            canvas.itemconfigure(errorproducts_text, text= str(num_errors) + " errors")
            
            while num_errors > 0 and running_cnt >= 0:
                print(array_errors)
                time.sleep(300)
                scan_errors(sheet)
                running_cnt = running_cnt -1
                workbook.save(save_filename)

            try:
                # Display a pop-up notification and play audio
                pygame.mixer.init()
                pygame.mixer.music.load(relative_to_assets('notification.mp3'))
                pygame.mixer.music.play()
            except:
                print("audio error")

            messagebox_text = scanning_text + " finished" +"\n " + str(num_errors) + " products failed to scan, please to check"
            # Display a pop-up notification
            win32api.MessageBox(
                None,
                messagebox_text,
                "Notification",
                win32con.MB_OK | win32con.MB_ICONINFORMATION
            )

            canvas.itemconfigure(scanning_status_text, text="finished")
            
            remaining_time = wait_time
            update_timer()
            for _ in range(wait_time):
                if stop_event.is_set():
                    break
                time.sleep(1)

    canvas.itemconfigure(scanning_id_text, text="Scanning Stopped")
    canvas.itemconfigure(scanning_status_text, text="")
    canvas.itemconfigure(remainingtime_text, text="")
    start_btn["state"] = "normal"

def start_processing():
    # Get the current date and time
    current_datetime = datetime.now()

    canvas.itemconfigure(start_time_text, text=current_datetime)
    
    global address1_value, address2_value, address3_value
    global city1_value, city2_value, city3_value
    global state1_value, state2_value, state3_value
    global telephone1_value, telephone2_value, telephone3_value
    global zipCode1_value, zipCode2_value, zipCode3_value
    global wait_time
    address1_value = address1_entry.get()
    city1_value = city1_entry.get()
    state1_value = state1_entry.get()
    telephone1_value = telephone1_entry.get()
    zipCode1_value = zipCode1_entry.get()
    address2_value = address2_entry.get()
    city2_value = city2_entry.get()
    state2_value = state2_entry.get()
    telephone2_value = telephone2_entry.get()
    zipCode2_value = zipCode2_entry.get()
    address3_value = address3_entry.get()
    city3_value = city3_entry.get()
    state3_value = state3_entry.get()
    telephone3_value = telephone3_entry.get()
    zipCode3_value = zipCode3_entry.get()

    try:
        wait_time = int(time_entry.get())
    except:
        wait_time = 0

    stop_event.clear()
    # Define file types to allow only XLSX files
    filetypes = [("Excel Files", "*.xlsx")]
    # Open the file dialog, restricted to XLSX files
    file_path = askopenfilename(filetypes=filetypes)

    if(file_path != ''):
        directory_path = os.path.dirname(file_path)

        workbook = openpyxl.load_workbook(file_path)

        sheet = workbook['Sheet1']

        # # Initial xlsx file
        # for column_letter in sheet.iter_cols(min_row=2):
        #     for cell in column_letter:
        #         if cell.column_letter != 'A':
        #             cell.value = None

        start_btn["state"] = "disabled"
        stop_btn["state"] = "normal"

        thread = threading.Thread(target = main_process, args=(workbook, sheet, directory_path))
        thread.start()
  
        
def stop_processing():
    stop_btn["state"] = "disabled"
    stop_event.set()
    

def handle_btn_press(option):
    if option=="start":
        start_processing()

    elif option=="stop":
        stop_processing()
        


# GUI starts here : 

# window.iconbitmap(relative_to_assets("beanonymous.ico"))

canvas.place(x = 0, y = 0)

# entry_image_1 = PhotoImage(
#     file=relative_to_assets("entry_1.png"))
# entry_bg_1 = canvas.create_image(
#     0,
#     90.0,
#     image=entry_image_1
# )
canvas.create_text(
    230.0,
    50.0,
    anchor="nw",
    text="Address",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    440.0,
    50.0,
    anchor="nw",
    text="City",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)
canvas.create_text(
    535.0,
    50.0,
    anchor="nw",
    text="State",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)
canvas.create_text(
    585.0,
    50.0,
    anchor="nw",
    text="zipCode",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)
canvas.create_text(
    670.0,
    50.0,
    anchor="nw",
    text="Telephone",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

address1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
address1_entry.place(
    x=145.0,
    y=84,
    width=220.0,
    height=30.0
)

city1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
city1_entry.place(
    x=375.0,
    y=84,
    width=150.0,
    height=30.0
)

state1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
state1_entry.place(
    x=535.0,
    y=84,
    width=40.0,
    height=30.0
)

zipCode1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
zipCode1_entry.place(
    x=585.0,
    y=84,
    width=40.0,
    height=30.0
)

telephone1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
telephone1_entry.place(
    x=635.0,
    y=84,
    width=150.0,
    height=30.0
)

canvas.create_text(
    20.0,
    88.0,
    anchor="nw",
    text="Testing Address 1",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

address2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
address2_entry.place(
    x=145.0,
    y=139,
    width=220.0,
    height=30.0
)

city2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
city2_entry.place(
    x=375.0,
    y=139,
    width=150.0,
    height=30.0
)

state2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
state2_entry.place(
    x=535.0,
    y=139,
    width=40.0,
    height=30.0
)

zipCode2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
zipCode2_entry.place(
    x=585.0,
    y=139,
    width=40.0,
    height=30.0
)

telephone2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
telephone2_entry.place(
    x=635.0,
    y=139,
    width=150.0,
    height=30.0
)

canvas.create_text(
    20.0,
    143.0,
    anchor="nw",
    text="Testing Address 2",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

address3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
address3_entry.place(
    x=145.0,
    y=194,
    width=220.0,
    height=30.0
)

city3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
city3_entry.place(
    x=375.0,
    y=194,
    width=150.0,
    height=30.0
)

state3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
state3_entry.place(
    x=535.0,
    y=194,
    width=40.0,
    height=30.0
)
zipCode3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
zipCode3_entry.place(
    x=585.0,
    y=194,
    width=40.0,
    height=30.0
)

telephone3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
telephone3_entry.place(
    x=635.0,
    y=194,
    width=150.0,
    height=30.0
)

canvas.create_text(
    20.0,
    198.0,
    anchor="nw",
    text="Testing Address 3",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

scanning_id_text = canvas.create_text(
    163.0,
    320.0,
    anchor="nw",
    text="",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

scanning_status_text = canvas.create_text(
    283.0,
    320.0,
    anchor="nw",
    text="",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

remainingtime_text = canvas.create_text(
    393.0,
    320.0,
    anchor="nw",
    text="",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

errorproducts_text = canvas.create_text(
    483.0,
    320.0,
    anchor="nw",
    text="",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    20.0,
    258.0,
    anchor="nw",
    text="Time to wait/halt",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

time_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
time_entry.place(
    x=145.0,
    y=254,
    width=80.0,
    height=30.0
)

canvas.create_text(
    350.0,
    258.0,
    anchor="nw",
    text="The Number of Threads",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

threadscnt_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
threadscnt_entry.place(
    x=525.0,
    y=254,
    width=80.0,
    height=30.0
)

canvas.create_text(
    40.0,
    370.0,
    anchor="nw",
    text="Thread 1",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    163.0,
    370.0,
    anchor="nw",
    text="Product ID",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    353.0,
    370.0,
    anchor="nw",
    text="Product URL",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

productID1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productID1_entry.place(
    x=245.0,
    y=365,
    width=80.0,
    height=30.0
)

productURL1_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productURL1_entry.place(
    x=445.0,
    y=365,
    width=300.0,
    height=30.0
)

canvas.create_text(
    40.0,
    420.0,
    anchor="nw",
    text="Thread 2",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    163.0,
    420.0,
    anchor="nw",
    text="Product ID",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    353.0,
    420.0,
    anchor="nw",
    text="Product URL",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

productID2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productID2_entry.place(
    x=245.0,
    y=420,
    width=80.0,
    height=30.0
)

productURL2_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productURL2_entry.place(
    x=445.0,
    y=420,
    width=300.0,
    height=30.0
)

canvas.create_text(
    40.0,
    475.0,
    anchor="nw",
    text="Thread 3",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    163.0,
    475.0,
    anchor="nw",
    text="Product ID",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    353.0,
    475.0,
    anchor="nw",
    text="Product URL",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

productID3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productID3_entry.place(
    x=245.0,
    y=475,
    width=80.0,
    height=30.0
)

productURL3_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productURL3_entry.place(
    x=445.0,
    y=475,
    width=300.0,
    height=30.0
)

canvas.create_text(
    40.0,
    530.0,
    anchor="nw",
    text="Thread 4",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    163.0,
    530.0,
    anchor="nw",
    text="Product ID",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    353.0,
    530.0,
    anchor="nw",
    text="Product URL",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

productID4_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productID4_entry.place(
    x=245.0,
    y=530,
    width=80.0,
    height=30.0
)

productURL4_entry = Entry(
    bd=0,
    bg="#2D2D2D",
    fg="#FFFFFF",
    highlightthickness=0
)
productURL4_entry.place(
    x=445.0,
    y=530,
    width=300.0,
    height=30.0
)

canvas.create_text(
    40.0,
    580.0,
    anchor="nw",
    text="Start Time",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

canvas.create_text(
    40.0,
    580.0,
    anchor="nw",
    text="Start Time",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

start_time_text = canvas.create_text(
    163.0,
    580.0,
    anchor="nw",
    text="",
    fill="#FFFFFF",
    font=("Roboto Medium", 14 * -1)
)

start_img = PhotoImage(file=relative_to_assets("start.png"))
start_btn = Button(
    image=start_img, borderwidth=0, highlightthickness=0, relief="flat", command=lambda : handle_btn_press("start"),activebackground= "#202020")
start_btn.place(x=158, y=620, width=172, height=47)

stop_img = PhotoImage(file=relative_to_assets("stop.png"))
stop_btn = Button(image=stop_img, borderwidth=0, highlightthickness=0, relief="flat", command=lambda : handle_btn_press("stop"),activebackground= "#202020", state="disabled")
stop_btn.place(x=488, y=620, width=172, height=47)

address1_entry.insert(0, "1632 Scholar Dr")
city1_entry.insert(0, "Lawrenceville")
state1_entry.insert(0, "GA")
zipCode1_entry.insert(0, "30043")
telephone1_entry.insert(0, "202-337-1412")

address2_entry.insert(0, "605 Saint Lawrence Blvd")
city2_entry.insert(0, "Eastlake")
state2_entry.insert(0, "OH")
zipCode2_entry.insert(0, "44095")
telephone2_entry.insert(0, "202-337-1412")

address3_entry.insert(0, "203 Magnolia St")
city3_entry.insert(0, "Siloam Springs")
state3_entry.insert(0, "AR")
zipCode3_entry.insert(0, "72761")
telephone3_entry.insert(0, "202-337-1412")

time_entry.insert(0, "0")
window.resizable(False, False)
window.mainloop()

# End of GUI Code